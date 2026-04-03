"""
reporter.py
-----------
Generates stakeholder-ready output from a ComparisonReport:

  1. HTML Report  → Visual, color-coded, Arabic RTL, per-law
  2. Excel Summary → One row per article, filterable, sortable

Both outputs land in the directories defined in .env
(REPORTS_DIR and SUMMARIES_DIR).
"""

import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from jinja2 import Environment, BaseLoader

from src.config import config
from src.comparator import ComparisonReport, MatchStatus, STATUS_EMOJI, STATUS_LABEL

logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────
# HTML Template
# ──────────────────────────────────────────────

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>تقرير المقارنة — {{ report.law_name }}</title>
  <style>
    /* ── Reset & Base ── */
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
    body {
      font-family: 'Segoe UI', Tahoma, Arial, sans-serif;
      background: #f0f2f5;
      color: #1a1a2e;
      direction: rtl;
    }

    /* ── Layout ── */
    .container  { max-width: 1400px; margin: 0 auto; padding: 24px; }

    /* ── Header ── */
    .header {
      background: linear-gradient(135deg, #1a1a2e 0%, #16213e 60%, #0f3460 100%);
      color: white;
      padding: 36px 40px;
      border-radius: 16px;
      margin-bottom: 24px;
      box-shadow: 0 8px 32px rgba(0,0,0,0.18);
    }
    .header h1 { font-size: 1.7rem; font-weight: 700; margin-bottom: 6px; }
    .header .sub { font-size: 0.95rem; opacity: 0.75; margin-bottom: 20px; }
    .header .meta-row {
      display: flex; gap: 32px; flex-wrap: wrap; margin-top: 16px;
    }
    .header .meta-item { font-size: 0.88rem; opacity: 0.85; }
    .header .meta-item span { font-weight: 600; opacity: 1; }

    /* ── Verdict Badge ── */
    .verdict {
      display: inline-block;
      padding: 6px 20px;
      border-radius: 20px;
      font-weight: 700;
      font-size: 1rem;
      margin-top: 14px;
    }
    .verdict-ممتاز      { background: #d4edda; color: #155724; }
    .verdict-جيد        { background: #cce5ff; color: #004085; }
    .verdict-مقبول      { background: #fff3cd; color: #856404; }
    .verdict-يحتاج-مراجعة { background: #f8d7da; color: #721c24; }

    /* ── Score Cards ── */
    .cards {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
      gap: 16px;
      margin-bottom: 24px;
    }
    .card {
      background: white;
      border-radius: 12px;
      padding: 20px 16px;
      text-align: center;
      box-shadow: 0 2px 12px rgba(0,0,0,0.07);
      border-top: 4px solid #ddd;
      transition: transform .15s;
    }
    .card:hover { transform: translateY(-3px); }
    .card .num  { font-size: 2rem; font-weight: 800; line-height: 1; }
    .card .lbl  { font-size: 0.8rem; color: #666; margin-top: 6px; }
    .card .pct  { font-size: 0.75rem; color: #999; margin-top: 2px; }
    .card-match      { border-color: #28a745; }
    .card-match .num { color: #28a745; }
    .card-near       { border-color: #ffc107; }
    .card-near .num  { color: #e6a800; }
    .card-mismatch   { border-color: #dc3545; }
    .card-mismatch .num { color: #dc3545; }
    .card-missing    { border-color: #6c757d; }
    .card-missing .num { color: #6c757d; }
    .card-extra      { border-color: #17a2b8; }
    .card-extra .num { color: #17a2b8; }
    .card-coverage   { border-color: #0f3460; }
    .card-coverage .num { color: #0f3460; }

    /* ── Progress Bar ── */
    .progress-section {
      background: white;
      border-radius: 12px;
      padding: 20px 24px;
      margin-bottom: 24px;
      box-shadow: 0 2px 12px rgba(0,0,0,0.07);
    }
    .progress-section h3 { font-size: 0.95rem; color: #444; margin-bottom: 14px; }
    .progress-bar {
      display: flex;
      height: 28px;
      border-radius: 8px;
      overflow: hidden;
      width: 100%;
    }
    .pb-match    { background: #28a745; }
    .pb-near     { background: #ffc107; }
    .pb-mismatch { background: #dc3545; }
    .pb-missing  { background: #adb5bd; }
    .pb-extra    { background: #17a2b8; }
    .progress-legend {
      display: flex; gap: 20px; flex-wrap: wrap; margin-top: 10px;
    }
    .legend-item {
      display: flex; align-items: center; gap: 6px;
      font-size: 0.82rem; color: #555;
    }
    .legend-dot {
      width: 12px; height: 12px; border-radius: 3px;
    }

    /* ── Metadata Card ── */
    .meta-card {
      background: white;
      border-radius: 12px;
      padding: 20px 24px;
      margin-bottom: 24px;
      box-shadow: 0 2px 12px rgba(0,0,0,0.07);
    }
    .meta-card h3 { font-size: 1rem; color: #333; margin-bottom: 14px;
                    border-bottom: 2px solid #f0f0f0; padding-bottom: 8px; }
    .meta-grid {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
      gap: 12px;
    }
    .meta-item-box {
      background: #f8f9fa;
      border-radius: 8px;
      padding: 12px 16px;
    }
    .meta-item-box .key  { font-size: 0.78rem; color: #888; margin-bottom: 4px; }
    .meta-item-box .val  { font-size: 0.95rem; font-weight: 600; color: #222; }
    .meta-match   { background: #d4edda; }
    .meta-match .val { color: #155724; }
    .meta-fail    { background: #f8d7da; }
    .meta-fail .val  { color: #721c24; }

    /* ── Filters ── */
    .filters {
      background: white;
      border-radius: 12px;
      padding: 16px 24px;
      margin-bottom: 20px;
      box-shadow: 0 2px 12px rgba(0,0,0,0.07);
      display: flex;
      gap: 12px;
      flex-wrap: wrap;
      align-items: center;
    }
    .filters label { font-size: 0.85rem; color: #555; font-weight: 600; }
    .filter-btn {
      padding: 7px 16px;
      border-radius: 20px;
      border: 2px solid #ddd;
      background: white;
      cursor: pointer;
      font-size: 0.82rem;
      font-family: inherit;
      transition: all .15s;
      direction: rtl;
    }
    .filter-btn:hover  { border-color: #0f3460; color: #0f3460; }
    .filter-btn.active { background: #0f3460; color: white; border-color: #0f3460; }
    .search-box {
      padding: 7px 14px;
      border-radius: 8px;
      border: 2px solid #ddd;
      font-size: 0.85rem;
      font-family: inherit;
      direction: rtl;
      min-width: 180px;
    }
    .search-box:focus { outline: none; border-color: #0f3460; }

    /* ── Articles Table ── */
    .table-section {
      background: white;
      border-radius: 12px;
      box-shadow: 0 2px 12px rgba(0,0,0,0.07);
      overflow: hidden;
      margin-bottom: 24px;
    }
    .table-header {
      padding: 16px 24px;
      border-bottom: 2px solid #f0f0f0;
      display: flex;
      justify-content: space-between;
      align-items: center;
    }
    .table-header h3 { font-size: 1rem; color: #333; }
    .table-header .count { font-size: 0.82rem; color: #888; }
    table {
      width: 100%;
      border-collapse: collapse;
      font-size: 0.85rem;
    }
    thead th {
      background: #f8f9fa;
      padding: 12px 16px;
      text-align: right;
      font-weight: 600;
      color: #555;
      border-bottom: 2px solid #eee;
      position: sticky;
      top: 0;
    }
    tbody tr { border-bottom: 2px solid #f0f0f0; transition: background .1s; }
    tbody tr:hover { background: #fafbff; }
    tbody td {
      padding: 14px 16px;
      vertical-align: top;
      line-height: 1.7;
    }
    td.art-num {
      font-weight: 700;
      color: #0f3460;
      white-space: nowrap;
      text-align: center;
      width: 70px;
    }
    td.score-cell { text-align: center; white-space: nowrap; width: 90px; }
    td.status-cell { text-align: center; width: 110px; }
    td.text-cell {
      max-width: 380px;
      min-width: 220px;
      white-space: pre-wrap;
      word-break: break-word;
      line-height: 1.8;
      vertical-align: top;
      padding: 14px 16px;
    }

    /* ── Status Badges ── */
    .badge {
      display: inline-block;
      padding: 3px 10px;
      border-radius: 12px;
      font-size: 0.76rem;
      font-weight: 600;
      white-space: nowrap;
    }
    .badge-MATCH      { background:#d4edda; color:#155724; }
    .badge-NEAR_MATCH { background:#fff3cd; color:#856404; }
    .badge-MISMATCH   { background:#f8d7da; color:#721c24; }
    .badge-MISSING    { background:#e2e3e5; color:#383d41; }
    .badge-EXTRA      { background:#d1ecf1; color:#0c5460; }

    /* ── Score Bar ── */
    .score-bar-wrap { display: flex; align-items: center; gap: 6px; }
    .score-bar {
      flex: 1; height: 6px; border-radius: 3px;
      background: #e9ecef; overflow: hidden; min-width: 50px;
    }
    .score-bar-fill { height: 100%; border-radius: 3px; }

    /* ── Diff Hint ── */
    .diff-hint {
      font-size: 0.75rem;
      color: #888;
      margin-top: 4px;
      font-family: monospace;
      background: #f8f8f8;
      padding: 4px 8px;
      border-radius: 4px;
      border-right: 3px solid #ffc107;
      display: block;
      white-space: pre-wrap;
      word-break: break-all;
    }


    /* ── Footer ── */
    .footer {
      text-align: center;
      color: #aaa;
      font-size: 0.78rem;
      padding: 20px;
    }

    /* ── Print ── */
    @media print {
      .filters { display: none; }
      body { background: white; }
      .card:hover { transform: none; }
    }
  </style>
</head>
<body>
<div class="container">

  <!-- Header -->
  <div class="header">
    <h1>تقرير مقارنة النصوص القانونية</h1>
    {% if company_name %}
    <div class="sub" style="margin-top:6px; font-size:0.9rem; opacity:0.8;">{{ company_name }}</div>
    {% endif %}
    <div class="sub">{{ report.law_name }}</div>
    <div class="meta-row">
      <div class="meta-item">رقم القانون: <span>{{ report.law_number }}/{{ report.year }}</span></div>
      <div class="meta-item">رقم المجلة: <span>{{ report.metadata.json_magazine }}</span></div>
      <div class="meta-item">تاريخ التقرير: <span>{{ generated_at }}</span></div>
    </div>
    <div class="verdict verdict-{{ report.overall_verdict | replace(' ', '-') }}">
      الحكم الإجمالي: {{ report.overall_verdict }}
    </div>
  </div>

  <!-- Score Cards -->
  <div class="cards">
    <div class="card card-coverage">
      <div class="num">{{ "%.1f"|format(report.coverage_pct) }}%</div>
      <div class="lbl">نسبة التغطية</div>
      <div class="pct">{{ report.total_json - report.count_missing }}/{{ report.total_json }} مادة</div>
    </div>
    <div class="card card-coverage">
      <div class="num">{{ "%.1f"|format(report.match_pct) }}%</div>
      <div class="lbl">نسبة التطابق</div>
      <div class="pct">من المواد المقارنة</div>
    </div>
    <div class="card card-match">
      <div class="num">{{ report.count_match }}</div>
      <div class="lbl">✅ تطابق تام</div>
    </div>
    <div class="card card-near">
      <div class="num">{{ report.count_near_match }}</div>
      <div class="lbl">⚠️ تطابق جزئي</div>
    </div>
    <div class="card card-mismatch">
      <div class="num">{{ report.count_mismatch }}</div>
      <div class="lbl">❌ تعارض</div>
    </div>
    <div class="card card-missing">
      <div class="num">{{ report.count_missing }}</div>
      <div class="lbl">🔍 غائب</div>
    </div>
    <div class="card card-extra">
      <div class="num">{{ report.count_extra }}</div>
      <div class="lbl">➕ زائد</div>
    </div>
  </div>

  <!-- Progress Bar -->
  <div class="progress-section">
    <h3>توزيع نتائج المقارنة</h3>
    <div class="progress-bar">
      <div class="pb-match"    style="width:{{ match_pct_bar }}%"    title="تطابق"></div>
      <div class="pb-near"     style="width:{{ near_pct_bar }}%"     title="تطابق جزئي"></div>
      <div class="pb-mismatch" style="width:{{ mismatch_pct_bar }}%" title="تعارض"></div>
      <div class="pb-missing"  style="width:{{ missing_pct_bar }}%"  title="غائب"></div>
      <div class="pb-extra"    style="width:{{ extra_pct_bar }}%"    title="زائد"></div>
    </div>
    <div class="progress-legend">
      <div class="legend-item"><div class="legend-dot" style="background:#28a745"></div> تطابق تام ({{ report.count_match }})</div>
      <div class="legend-item"><div class="legend-dot" style="background:#ffc107"></div> تطابق جزئي ({{ report.count_near_match }})</div>
      <div class="legend-item"><div class="legend-dot" style="background:#dc3545"></div> تعارض ({{ report.count_mismatch }})</div>
      <div class="legend-item"><div class="legend-dot" style="background:#adb5bd"></div> غائب ({{ report.count_missing }})</div>
      <div class="legend-item"><div class="legend-dot" style="background:#17a2b8"></div> زائد ({{ report.count_extra }})</div>
    </div>
  </div>

  <!-- Metadata -->
  <div class="meta-card">
    <h3>مقارنة البيانات الوصفية</h3>
    <div class="meta-grid">
      <div class="meta-item-box {{ 'meta-match' if report.metadata.match else 'meta-fail' }}">
        <div class="key">رقم المجلة</div>
        <div class="val">
          {{ 'متطابق ✅' if report.metadata.match else 'غير متطابق ❌' }}
          — المصدر 1: {{ report.metadata.json_magazine }}
          | المصدر 2: {{ report.metadata.txt_magazine }}
        </div>
      </div>
    </div>
  </div>

  <!-- Filters -->
  <div class="filters">
    <label>تصفية حسب:</label>
    <button class="filter-btn active" onclick="filterTable('ALL')">الكل ({{ report.articles|length }})</button>
    <button class="filter-btn" onclick="filterTable('MATCH')">✅ تطابق ({{ report.count_match }})</button>
    <button class="filter-btn" onclick="filterTable('NEAR_MATCH')">⚠️ جزئي ({{ report.count_near_match }})</button>
    <button class="filter-btn" onclick="filterTable('MISMATCH')">❌ تعارض ({{ report.count_mismatch }})</button>
    <button class="filter-btn" onclick="filterTable('MISSING')">🔍 غائب ({{ report.count_missing }})</button>
    <button class="filter-btn" onclick="filterTable('EXTRA')">➕ زائد ({{ report.count_extra }})</button>
    <input class="search-box" type="text" placeholder="بحث برقم المادة..." oninput="searchTable(this.value)" />
  </div>

  <!-- Articles Table -->
  <div class="table-section">
    <div class="table-header">
      <h3>تفاصيل المقارنة</h3>
      <span class="count" id="visible-count">{{ report.articles|length }} مادة</span>
    </div>
    <table id="articles-table">
      <colgroup>
        <col style="width: 80px;">
        <col style="width: 120px;">
        <col style="width: 130px;">
        <col style="width: calc(50% - 165px);">
        <col style="width: calc(50% - 165px);">
      </colgroup>
      <thead>
        <tr>
          <th style="text-align:center;">رقم المادة</th>
          <th style="text-align:center;">الحالة</th>
          <th style="text-align:center;">نسبة التطابق</th>
          <th>نص المصدر الأول (JSON)</th>
          <th>نص المصدر الثاني (TXT)</th>
        </tr>
      </thead>
      <tbody>
        {% for art in report.articles %}
        <tr data-status="{{ art.status.value }}" data-num="{{ art.article_number }}">
          <td class="art-num">{{ art.article_number }}</td>
          <td class="status-cell">
            <span class="badge badge-{{ art.status.value }}">
              {{ status_emoji[art.status] }} {{ status_label[art.status] }}
            </span>
          </td>
          <td class="score-cell">
            {% if art.status not in ['MISSING', 'EXTRA'] %}
            <div class="score-bar-wrap">
              <div class="score-bar">
                <div class="score-bar-fill"
                     style="width:{{ art.similarity_score }}%;
                            background:{{ '#28a745' if art.similarity_score >= 95
                                     else '#ffc107' if art.similarity_score >= 80
                                     else '#dc3545' }}">
                </div>
              </div>
              <span>{{ "%.0f"|format(art.similarity_score) }}%</span>
            </div>
            {% else %}
            <span style="color:#ccc">—</span>
            {% endif %}
          </td>
          <td class="text-cell" id="j-{{ loop.index }}">
            {{ art.json_text if art.json_text else '—' }}
            {% if art.diff_hint %}
            <span class="diff-hint">{{ art.diff_hint }}</span>
            {% endif %}
          </td>
          <td class="text-cell" id="t-{{ loop.index }}">
            {{ art.txt_text if art.txt_text else '—' }}
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>

  <div class="footer">
    تم إنشاء هذا التقرير تلقائياً بواسطة نظام مقارنة النصوص القانونية
    {% if company_name %}— {{ company_name }}{% endif %}
    | {{ generated_at }}
  </div>
</div>

<script>
  let currentFilter = 'ALL';

  function filterTable(status) {
    currentFilter = status;
    document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
    event.target.classList.add('active');
    applyFilters();
  }

  function searchTable(val) {
    applyFilters(val.trim());
  }

  function applyFilters(search = '') {
    const rows = document.querySelectorAll('#articles-table tbody tr');
    let visible = 0;
    rows.forEach(row => {
      const statusMatch = currentFilter === 'ALL' || row.dataset.status === currentFilter;
      const searchMatch = !search || row.dataset.num.includes(search);
      const show = statusMatch && searchMatch;
      row.style.display = show ? '' : 'none';
      if (show) visible++;
    });
    document.getElementById('visible-count').textContent = visible + ' مادة';
  }


</script>
</body>
</html>"""


# ──────────────────────────────────────────────
# HTML Reporter
# ──────────────────────────────────────────────

def _generate_html(report: ComparisonReport, output_path: Path) -> None:
    """Render the HTML report from the Jinja2 template."""

    total = report.total_json + report.count_extra or 1

    env      = Environment(loader=BaseLoader())
    template = env.from_string(HTML_TEMPLATE)

    html = template.render(
        report          = report,
        generated_at    = datetime.now().strftime("%Y-%m-%d %H:%M"),
        company_name    = config.REPORT_COMPANY_NAME,
        status_emoji    = STATUS_EMOJI,
        status_label    = STATUS_LABEL,
        match_pct_bar   = round(report.count_match      / total * 100, 1),
        near_pct_bar    = round(report.count_near_match / total * 100, 1),
        mismatch_pct_bar= round(report.count_mismatch   / total * 100, 1),
        missing_pct_bar = round(report.count_missing    / total * 100, 1),
        extra_pct_bar   = round(report.count_extra      / total * 100, 1),
    )

    output_path.write_text(html, encoding="utf-8")
    logger.info(f"HTML report saved: {output_path}")


# ──────────────────────────────────────────────
# Excel Reporter
# ──────────────────────────────────────────────

def _generate_excel(report: ComparisonReport, output_path: Path) -> None:
    """Generate a structured Excel workbook with two sheets."""

    # ── Sheet 1: Summary ──────────────────────
    summary_data = {
        "البند":   [
            "اسم القانون", "رقم القانون", "السنة",
            "رقم المجلة (JSON)", "رقم المجلة (TXT)", "تطابق المجلة",
            "إجمالي مواد JSON", "إجمالي مواد TXT",
            "نسبة التغطية", "نسبة التطابق", "الحكم الإجمالي",
            "تطابق تام", "تطابق جزئي", "تعارض", "غائب", "زائد",
        ],
        "القيمة": [
            report.law_name, report.law_number, report.year,
            report.metadata.json_magazine,
            report.metadata.txt_magazine,
            "نعم ✅" if report.metadata.match else "لا ❌",
            report.total_json, report.total_txt,
            f"{report.coverage_pct:.1f}%",
            f"{report.match_pct:.1f}%",
            report.overall_verdict,
            report.count_match,
            report.count_near_match,
            report.count_mismatch,
            report.count_missing,
            report.count_extra,
        ],
    }

    # ── Sheet 2: Article details ──────────────
    rows = []
    for art in report.articles:
        rows.append({
            "رقم المادة":       art.article_number,
            "الحالة":           STATUS_LABEL[art.status],
            "رمز الحالة":       STATUS_EMOJI[art.status],
            "نسبة التطابق %":   art.similarity_score if art.similarity_score > 0 else "",
            "نص JSON":          art.json_text,
            "نص TXT":           art.txt_text,
            "ملاحظة الفارق":    art.diff_hint,
        })

    df_summary = pd.DataFrame(summary_data)
    df_articles = pd.DataFrame(rows)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="ملخص", index=False)
        df_articles.to_excel(writer, sheet_name="تفاصيل المواد", index=False)

        # ── Style the sheets ──────────────────
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )

        STATUS_COLORS = {
            "تطابق":               "D4EDDA",
            "تطابق جزئي":          "FFF3CD",
            "تعارض":               "F8D7DA",
            "غائب عن المصدر الثاني": "E2E3E5",
            "زائد في المصدر الثاني": "D1ECF1",
        }

        thin = Side(style="thin", color="DDDDDD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Style summary sheet
        ws_summary = writer.sheets["ملخص"]
        ws_summary.column_dimensions["A"].width = 28
        ws_summary.column_dimensions["B"].width = 45
        for row in ws_summary.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border    = border
            # Header row
            if row[0].row == 1:
                for cell in row:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="0F3460")

        # Style articles sheet
        ws_articles = writer.sheets["تفاصيل المواد"]
        ws_articles.column_dimensions["A"].width = 14
        ws_articles.column_dimensions["B"].width = 22
        ws_articles.column_dimensions["C"].width = 8
        ws_articles.column_dimensions["D"].width = 16
        ws_articles.column_dimensions["E"].width = 60
        ws_articles.column_dimensions["F"].width = 60
        ws_articles.column_dimensions["G"].width = 50

        for row in ws_articles.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="right", vertical="top", wrap_text=True
                )
                cell.border = border

            # Header row
            if row[0].row == 1:
                for cell in row:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="0F3460")
                continue

            # Color-code by status
            status_val = row[1].value
            color = STATUS_COLORS.get(status_val)
            if color:
                fill = PatternFill("solid", fgColor=color)
                for cell in row[:4]:   # color first 4 cols only
                    cell.fill = fill

    logger.info(f"Excel summary saved: {output_path}")


# ──────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────

def generate_report(report: ComparisonReport) -> dict[str, Path]:
    """
    Generate all output files for a ComparisonReport.

    Returns:
        dict with keys 'html' and 'excel' pointing to output paths.
    """
    config.REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    config.SUMMARIES_DIR.mkdir(parents=True, exist_ok=True)

    timestamp    = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name    = f"{report.law_id}_{timestamp}"

    html_path    = config.REPORTS_DIR  / f"{base_name}.html"
    excel_path   = config.SUMMARIES_DIR / f"{base_name}.xlsx"

    _generate_html(report,  html_path)
    _generate_excel(report, excel_path)

    logger.info(f"Reports generated for {report.law_id}")
    return {"html": html_path, "excel": excel_path}


# ──────────────────────────────────────────────
# Quick Self-Test
# ──────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    import argparse
    import logging

    logging.basicConfig(
        level  = logging.INFO,
        format = "%(levelname)-8s %(message)s"
    )

    parser = argparse.ArgumentParser(
        description="Reporter module — generate HTML + Excel reports."
    )
    parser.add_argument("--json",      type=str, required=True)
    parser.add_argument("--txt",       type=str, required=True)
    parser.add_argument("--law-index", type=int, default=0)
    args = parser.parse_args()

    from pathlib import Path
    from src.ingestion import load_pair
    from src.extractor import extract
    from src.comparator import compare

    # Load → Extract → Compare → Report
    pair = load_pair(args.json, args.txt, law_index=args.law_index)

    for enc in ["utf-8-sig", "utf-8", "cp1256"]:
        try:
            raw = Path(args.txt).read_text(encoding=enc)
            break
        except UnicodeDecodeError:
            continue

    extracted = extract(raw)
    report    = compare(pair.source1, extracted, pair.law_id)
    paths     = generate_report(report)

    SEP = "=" * 55
    print()
    print(SEP)
    print("  REPORTS GENERATED SUCCESSFULLY")
    print(SEP)
    print(f"  HTML   : {paths['html']}")
    print(f"  Excel  : {paths['excel']}")
    print(SEP)
    print()
    print("  Open the HTML file in your browser to view the report.")
    print()